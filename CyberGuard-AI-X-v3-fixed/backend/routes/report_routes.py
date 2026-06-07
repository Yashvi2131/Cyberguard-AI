"""
CyberGuard AI X - Report Generation Routes
PDF and Excel report export
"""

import io
import os
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from database.db import execute_query

report_bp = Blueprint('reports', __name__)


def generate_pdf_report(complaints: list, title: str = "CyberGuard AI X - Complaint Report") -> bytes:
    """Generate PDF report using reportlab (produces valid, openable PDFs)"""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, KeepTogether
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.5*cm, leftMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=2*cm,
            title="CyberGuard AI X Report",
            author="CyberGuard AI X"
        )

        # ── Color palette ──
        DARK_NAVY   = colors.HexColor('#0A0A1E')
        CYBER_BLUE  = colors.HexColor('#00B4FF')
        CYBER_GREEN = colors.HexColor('#00FF88')
        CYBER_RED   = colors.HexColor('#FF4444')
        CYBER_YEL   = colors.HexColor('#FFCC00')
        CYBER_PUR   = colors.HexColor('#AA66FF')
        MID_GRAY    = colors.HexColor('#444466')
        LIGHT_GRAY  = colors.HexColor('#F5F5FA')
        WHITE       = colors.white
        BLACK       = colors.HexColor('#111122')

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('CyberTitle',
            fontSize=22, fontName='Helvetica-Bold',
            textColor=DARK_NAVY, alignment=TA_CENTER, spaceAfter=4)
        sub_style = ParagraphStyle('CyberSub',
            fontSize=10, fontName='Helvetica',
            textColor=MID_GRAY, alignment=TA_CENTER, spaceAfter=2)
        section_style = ParagraphStyle('Section',
            fontSize=11, fontName='Helvetica-Bold',
            textColor=DARK_NAVY, spaceAfter=6, spaceBefore=12)
        cell_style = ParagraphStyle('Cell',
            fontSize=7.5, fontName='Helvetica',
            textColor=BLACK, leading=10)
        cell_bold = ParagraphStyle('CellBold',
            fontSize=7.5, fontName='Helvetica-Bold',
            textColor=BLACK, leading=10)
        footer_style = ParagraphStyle('Footer',
            fontSize=8, fontName='Helvetica-Oblique',
            textColor=MID_GRAY, alignment=TA_CENTER)

        story = []

        # ── Header banner ──
        story.append(Paragraph("🛡 CyberGuard AI X", title_style))
        story.append(Paragraph("Cybercrime Complaint Intelligence Report", sub_style))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}  |  Total Complaints: {len(complaints)}",
            sub_style))
        story.append(HRFlowable(width="100%", thickness=2, color=CYBER_BLUE, spaceAfter=10))

        # ── Summary statistics ──
        total  = len(complaints)
        high   = sum(1 for c in complaints if c.get('severity') == 'High')
        medium = sum(1 for c in complaints if c.get('severity') == 'Medium')
        low    = sum(1 for c in complaints if c.get('severity') == 'Low')
        pending = sum(1 for c in complaints if c.get('status') == 'Pending')
        resolved = sum(1 for c in complaints if c.get('status') == 'Resolved')

        # Category counts
        cat_counts = {}
        for c in complaints:
            cat = c.get('category', 'Unknown')
            cat_counts[cat] = cat_counts.get(cat, 0) + 1

        sum_data = [
            ['Metric', 'Value', 'Metric', 'Value'],
            ['Total Complaints', str(total), 'High Severity', str(high)],
            ['Medium Severity', str(medium), 'Low Severity', str(low)],
            ['Pending', str(pending), 'Resolved', str(resolved)],
        ]
        sum_table = Table(sum_data, colWidths=[5*cm, 3*cm, 5*cm, 3*cm])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), DARK_NAVY),
            ('TEXTCOLOR',  (0, 0), (-1, 0), CYBER_GREEN),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 9),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LIGHT_GRAY, WHITE]),
            ('FONTNAME',   (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME',   (2, 1), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR',  (1, 1), (1, -1), CYBER_BLUE),
            ('TEXTCOLOR',  (3, 1), (3, -1), CYBER_BLUE),
            ('FONTSIZE',   (0, 1), (-1, -1), 9),
            ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCDD')),
            ('ROUNDEDCORNERS', [4]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(Paragraph("Summary Statistics", section_style))
        story.append(sum_table)
        story.append(Spacer(1, 0.4*cm))

        # ── Category breakdown ──
        if cat_counts:
            cat_data = [['Category', 'Count', 'Share %']]
            for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
                pct = f"{cnt/total*100:.1f}%" if total else "0%"
                cat_data.append([cat, str(cnt), pct])
            cat_table = Table(cat_data, colWidths=[7*cm, 3*cm, 3*cm])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, 0), DARK_NAVY),
                ('TEXTCOLOR',     (0, 0), (-1, 0), CYBER_GREEN),
                ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0), (-1, -1), 9),
                ('ALIGN',         (1, 0), (-1, -1), 'CENTER'),
                ('ROWBACKGROUNDS',(0, 1), (-1, -1), [LIGHT_GRAY, WHITE]),
                ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCDD')),
                ('TOPPADDING',    (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(Paragraph("Category Breakdown", section_style))
            story.append(cat_table)
            story.append(Spacer(1, 0.4*cm))

        # ── Complaints detail table ──
        story.append(Paragraph("Complaint Details", section_style))

        def sev_color(sev):
            return {
                'High':   colors.HexColor('#FFDDDD'),
                'Medium': colors.HexColor('#FFF3CC'),
                'Low':    colors.HexColor('#DDFFF0'),
            }.get(sev, WHITE)

        col_widths = [1.2*cm, 7.5*cm, 3.8*cm, 2.5*cm, 2.8*cm, 4.5*cm, 2.4*cm, 2.3*cm]
        headers = ['ID', 'Complaint Text', 'Category', 'Severity', 'Status', 'Department', 'User', 'Date']
        table_data = [[Paragraph(h, ParagraphStyle('Hdr', fontName='Helvetica-Bold',
                        fontSize=8, textColor=WHITE, alignment=TA_CENTER)) for h in headers]]

        style_cmds = [
            ('BACKGROUND',    (0, 0), (-1, 0), DARK_NAVY),
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCDD')),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 4),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]

        for i, complaint in enumerate(complaints[:150]):
            sev = str(complaint.get('severity', 'Low'))
            txt = str(complaint.get('complaint_text', ''))
            if len(txt) > 90: txt = txt[:87] + '...'
            dept = str(complaint.get('department', ''))
            if len(dept) > 30: dept = dept[:27] + '...'
            date_val = str(complaint.get('created_at', ''))[:10] if complaint.get('created_at') else ''

            row = [
                Paragraph(str(complaint.get('id', '')),      cell_style),
                Paragraph(txt,                                cell_style),
                Paragraph(str(complaint.get('category', '')), cell_style),
                Paragraph(sev,                                cell_bold),
                Paragraph(str(complaint.get('status', '')),   cell_style),
                Paragraph(dept,                               cell_style),
                Paragraph(str(complaint.get('user_name', ''))[:18], cell_style),
                Paragraph(date_val,                           cell_style),
            ]
            table_data.append(row)
            row_idx = i + 1
            style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx),
                                LIGHT_GRAY if i % 2 == 0 else WHITE))
            style_cmds.append(('BACKGROUND', (3, row_idx), (3, row_idx), sev_color(sev)))

        detail_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        detail_table.setStyle(TableStyle(style_cmds))
        story.append(detail_table)

        story.append(Spacer(1, 0.6*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#CCCCDD')))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            "Confidential — CyberGuard AI X | Government Cybercrime Intelligence Platform | cybercrime.gov.in | Helpline: 1930",
            footer_style))

        doc.build(story)
        buffer.seek(0)
        return buffer.read()

    except ImportError as e:
        # Fallback minimal PDF if reportlab missing
        return (b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj "
                b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj "
                b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 595 842]"
                b"/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj "
                b"4 0 obj<</Length 44>>stream\nBT /F1 14 Tf 100 750 Td"
                b"(Install reportlab: pip install reportlab)Tj ET\nendstream endobj "
                b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj "
                b"xref\n0 6\n0000000000 65535 f\n"
                b"trailer<</Size 6/Root 1 0 R>>\nstartxref\n0\n%%EOF")


def generate_excel_report(complaints: list) -> bytes:
    """Generate Excel report using openpyxl"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Complaints Report"

        # Styles
        header_fill = PatternFill(start_color="0A0A1E", end_color="0A0A1E", fill_type="solid")
        header_font = Font(color="00FF88", bold=True, size=11)
        high_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
        medium_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
        low_fill = PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")

        # Title row
        ws.merge_cells('A1:I1')
        ws['A1'] = 'CyberGuard AI X - Complaint Intelligence Report'
        ws['A1'].font = Font(bold=True, size=14, color="000000")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:I2')
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Total: {len(complaints)}'
        ws['A2'].alignment = Alignment(horizontal='center')

        # Headers
        headers = ['ID', 'Complaint Text', 'Category', 'Confidence %', 'Severity', 'Department', 'Status', 'User', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, complaint in enumerate(complaints, 5):
            row_data = [
                complaint.get('id'),
                str(complaint.get('complaint_text', ''))[:100],
                complaint.get('category', ''),
                round(float(complaint.get('confidence_score', 0)), 2),
                complaint.get('severity', ''),
                complaint.get('department', ''),
                complaint.get('status', ''),
                complaint.get('user_name', ''),
                str(complaint.get('created_at', ''))[:10] if complaint.get('created_at') else ''
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Color severity
                if col_idx == 5:
                    if value == 'High':
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif value == 'Medium':
                        cell.fill = PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="CCFFE8", end_color="CCFFE8", fill_type="solid")

        # Auto-width columns
        col_widths_xl = [8, 50, 18, 14, 12, 30, 16, 20, 14]
        for i, width in enumerate(col_widths_xl, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2['A1'] = 'Category'
        ws2['B1'] = 'Count'
        ws2['A1'].font = Font(bold=True)
        ws2['B1'].font = Font(bold=True)

        category_counts = {}
        for c in complaints:
            cat = c.get('category', 'Unknown')
            category_counts[cat] = category_counts.get(cat, 0) + 1

        for row_idx, (cat, count) in enumerate(sorted(category_counts.items()), 2):
            ws2.cell(row=row_idx, column=1, value=cat)
            ws2.cell(row=row_idx, column=2, value=count)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    except ImportError:
        return b'Excel generation requires openpyxl. Install: pip install openpyxl'


@report_bp.route('/pdf', methods=['GET'])
@jwt_required()
def export_pdf():
    user_id = int(get_jwt_identity())
    db_user = execute_query("SELECT id, role FROM users WHERE id = %s", (user_id,), fetchone=True) or {}
    role = db_user.get('role', 'user')
    uid = db_user.get('id', user_id)

    where = "" if role == 'admin' else f"WHERE c.user_id = {uid}"
    complaints = execute_query(
        f"""SELECT c.*, u.name as user_name FROM complaints c
            JOIN users u ON c.user_id = u.id {where}
            ORDER BY c.created_at DESC LIMIT 200""",
        fetchall=True
    ) or []

    pdf_bytes = generate_pdf_report(complaints)
    filename = f"cyberguard_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )


@report_bp.route('/excel', methods=['GET'])
@jwt_required()
def export_excel():
    user_id = int(get_jwt_identity())
    db_user = execute_query("SELECT id, role FROM users WHERE id = %s", (user_id,), fetchone=True) or {}
    role = db_user.get('role', 'user')
    uid = db_user.get('id', user_id)

    where = "" if role == 'admin' else f"WHERE c.user_id = {uid}"
    complaints = execute_query(
        f"""SELECT c.*, u.name as user_name FROM complaints c
            JOIN users u ON c.user_id = u.id {where}
            ORDER BY c.created_at DESC LIMIT 500""",
        fetchall=True
    ) or []

    excel_bytes = generate_excel_report(complaints)
    filename = f"cyberguard_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        io.BytesIO(excel_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
